# -*- coding: utf-8 -*-
print("module [backend.api_project] loaded")


from backend import check_token, ProcessingException
from backend_model.table_project import *
from backend_model.table_setup import SetupMenu
from backend import manager
from backend_lib.lib_project import *
from flask import make_response, jsonify, request, json, send_file, render_template, abort, send_from_directory
from flask_restful import reqparse
from werkzeug.utils import secure_filename
from backend import app
import os
from urllib.parse import urlparse, quote
from uuid import uuid4
import shutil
db = DBManager.db
from datetime import datetime
# REST API(s) available :
#  - /api/v1/login
#  - /api/v1/logout


def folders_in(path_to_parent):
    for fname in os.listdir(path_to_parent):
        if os.path.isdir(os.path.join(path_to_parent, fname)):
            yield os.path.join(path_to_parent, fname)


def has_folders(path_to_parent):
    folders = list(folders_in(path_to_parent))
    return len(folders) != 0


@app.route("/api/mes/v1/project/file_manager", methods=["GET"])
def filemanager_get_api():
    parameter_dict = request.args.to_dict()
    if len(parameter_dict) == 0:
        return make_response(jsonify('No Parameters'), 400)

    BASE_DIR = app.config['UPLOAD_BASE_DIR']

    if parameter_dict['command'] == "GetDirContents":
        arguments_dict = json.loads(parameter_dict['arguments'])

        abs_path = os.path.join(BASE_DIR, "")
        for dir_item in arguments_dict['pathInfo']:
            abs_path = os.path.join(abs_path, dir_item['key'])

        if not os.path.exists(abs_path):
            return make_response(jsonify('Path is invalid'), 404)

        files = os.listdir(abs_path)

        result = []

        for file in files:
            file_item = dict()
            file_path = os.path.join(abs_path, file)
            file_item['key'] = file_item['name'] = file

            file_item['dateModified'] = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%dT%H:%M:%S')
            file_item['isDirectory'] = os.path.isdir(file_path)
            file_item['size'] = os.path.getsize(file_path)
            if file_item['isDirectory'] is True:
                file_item['hasSubDirectories'] = has_folders(file_path)
            else:
                file_item['hasSubDirectories'] = False
            result.append(file_item)

        ret_json = dict()
        ret_json['success'] = True
        ret_json['errorCode'] = None
        ret_json['errorText'] = ""
        ret_json['result'] = result

    return make_response(jsonify(ret_json), 200)


@app.route("/api/mes/v1/project/file_manager", methods=["POST"])
def filemanager_post_api():
    parameter_dict = request.args.to_dict()

    if len(parameter_dict) == 0:
        parameter_dict = request.form.to_dict()
        if len(parameter_dict) == 0:
            return make_response(jsonify('No Parameters'), 400)

    BASE_DIR = app.config['UPLOAD_BASE_DIR']

    arguments_dict = json.loads(parameter_dict['arguments'])
    abs_path = os.path.join(BASE_DIR, "")

    if parameter_dict['command'] == "CreateDir":
        for dir_item in arguments_dict['pathInfo']:
            abs_path = os.path.join(abs_path, dir_item['key'])

        abs_path = os.path.join(abs_path, arguments_dict['name'])
        os.mkdir(abs_path)
    elif parameter_dict['command'] == "Remove":
        for dir_item in arguments_dict['pathInfo']:
            abs_path = os.path.join(abs_path, dir_item['key'])

        if arguments_dict['isDirectory'] is True:
            os.rmdir(abs_path)
        else:
            os.remove(abs_path)
    elif parameter_dict['command'] == "Rename":
        for dir_item in arguments_dict['pathInfo']:
            abs_path = os.path.join(abs_path, dir_item['key'])

        new_path = os.path.join(os.path.dirname(abs_path), arguments_dict['name'])
        os.rename(abs_path, new_path)
    elif parameter_dict['command'] == "Copy":
        src_path = abs_path
        for dir_item in arguments_dict['sourcePathInfo']:
            src_path = os.path.join(src_path, dir_item['key'])

        dst_path = abs_path
        for dir_item in arguments_dict['destinationPathInfo']:
            dst_path = os.path.join(dst_path, dir_item['key'])

        if arguments_dict['sourceIsDirectory'] is True:
            shutil.copytree(src_path, dst_path)
        else:
            dst_path = os.path.join(dst_path, os.path.basename(src_path))
            shutil.copy(src_path, dst_path)
    elif parameter_dict['command'] == "Move":
        src_path = abs_path
        for dir_item in arguments_dict['sourcePathInfo']:
            src_path = os.path.join(src_path, dir_item['key'])

        dst_path = abs_path
        for dir_item in arguments_dict['destinationPathInfo']:
            dst_path = os.path.join(dst_path, dir_item['key'])

        if arguments_dict['sourceIsDirectory'] is True:
            shutil.move(src_path, dst_path)
        else:
            dst_path = os.path.join(dst_path, os.path.basename(src_path))
            shutil.move(src_path, dst_path)
    elif parameter_dict['command'] == "Download":
        down_path = abs_path
        for pathinfo_item in arguments_dict['pathInfoList']:
            for dir_item in pathinfo_item:
                down_path = os.path.join(down_path, dir_item['key'])

        return send_file(down_path,
                         mimetype='application/octet-stream',
                         attachment_filename=os.path.basename(down_path),
                         as_attachment=True)
    elif parameter_dict['command'] == 'UploadChunk':
        metadata = json.loads(arguments_dict['chunkMetadata'])

        dst_path = abs_path
        for dir_item in arguments_dict['destinationPathInfo']:
            dst_path = os.path.join(dst_path, dir_item['key'])

        fo = request.files['chunk']
        upload_path = os.path.join(dst_path, metadata['UploadId'])

        with open(upload_path, 'ab') as f:
            offset = metadata['FileSize'] * metadata['Index']
            f.seek(offset)
            f.write(fo.stream.read())
            if metadata['TotalCount'] - 1 == metadata['Index']:
                rename_path = os.path.join(dst_path, metadata['FileName'])
                os.rename(upload_path, rename_path)

    ret_json = dict()
    ret_json['success'] = True
    ret_json['errorCode'] = None
    ret_json['errorText'] = ""
    ret_json['result'] = ""

    return make_response(jsonify(ret_json), 200)

@app.route('/api/mes/v1/project-excution-plan-subcontract/upload', methods=['POST'])
def upload_excution_plan_subcontract():
    f = request.files['file']
    if f is None:
        return make_response('require parameter is missing', 400)
    filename = str(uuid4())[:8] + '__' + f.filename
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], "project-excution-plan-subcontract")
    abs_path = os.path.join(base_path, filename)

    if not os.path.exists(base_path):
        os.makedirs(base_path)

    f.save(abs_path)
    return make_response(filename, 200)

@app.route('/api/mes/v1/project-excution-plan-subcontract/<filepath>/<filename>', methods=['GET'])
def download_excution_plan_subcontract(filepath, filename):
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], "project-excution-plan-subcontract")
    try:
        encoded_filename = quote(filename.encode('utf-8'))
        r = send_from_directory(base_path, filename=filepath, as_attachment=True)
        r.headers.set('Content-Disposition', f'attachment; filename={encoded_filename}')
        return r
    except FileNotFoundError:
        return make_response('file not found', 404)

@app.route('/api/mes/v1/project-attachment/document', methods=['POST'])
def upload_document():
    f = request.files['file']
    if f is None:
        return make_response('require parameter is missing', 400)
    filename = str(uuid4())[:8] + '__' + f.filename
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], "project-document")
    abs_path = os.path.join(base_path, filename)

    if not os.path.exists(base_path):
        os.makedirs(base_path)

    f.save(abs_path)
    return make_response(filename, 200)

@app.route('/api/mes/v1/project-attachment/daily-log', methods=['POST'])
def upload_daily_log():
    f = request.files['file']
    if f is None:
        return make_response('require parameter is missing', 400)
    filename = str(uuid4())[:8] + '__' + f.filename
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], "project-daily-log")
    abs_path = os.path.join(base_path, filename)

    if not os.path.exists(base_path):
        os.makedirs(base_path)

    f.save(abs_path)
    return make_response(filename, 200)

@app.route('/api/mes/v1/project-document/<filename>', methods=['GET'])
def download_document(filename):
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], 'project-document')
    try:
        encoded_filename = quote(filename.encode('utf-8'))
        r = send_from_directory(base_path, filename=filename, as_attachment=True)
        r.headers.set('Contetn-Disposition', f'attachment; filename={encoded_filename}')
        return r
    except FileNotFoundError:
        return make_response('file not found', 404)

@app.route('/api/mes/v1/project-daily-log/<filename>', methods=['GET'])
def download_daily_log(filename):
    base_path = os.path.join(app.config['UPLOAD_BASE_DIR'], 'project-daily-log')
    try:
        encoded_filename = quote(filename.encode('utf-8'))
        r = send_from_directory(base_path, filename=filename, as_attachment=True)
        r.headers.set('Contetn-Disposition', f'attachment; filename={encoded_filename}')
        return r
    except FileNotFoundError:
        return make_response('file not found', 404)

@app.route('/api/server/v1/project-document/remove/<int:document_id>', methods=['POST'])
def project_document_remove(document_id):
    document = db.session.query(ProjectDocument).filter(ProjectDocument.id == document_id).first()
    file_path = os.path.join(app.config['UPLOAD_BASE_DIR'], document.file_path)
    try:
         os.remove(file_path)
    except IsADirectoryError:
        pass

    return make_response('success', 200)

@app.route('/api/server/v1/project-daily-log/remove/<int:daily_log_id>', methods=['POST'])
def project_daily_log_remove(daily_log_id):
    daily_log = db.session.query(ProjectDailyLog).filter(ProjectDailyLog.id == daily_log_id).first()
    file_path = os.path.join(app.config['UPLOAD_BASE_DIR'], daily_log.attachment_path)
    try:
        os.remove(file_path)
    except IsADirectoryError:
        pass

    return make_response('success', 200)    

@app.route("/api/mes/v1/project/registration-to-work-order", methods=["POST"])
def registration_to_work_order():
    parser = reqparse.RequestParser()
    parser.add_argument("project_management_id", type=int, location="json")
    parser.add_argument("department", type=str, location="json")
    parser.add_argument("manager", type=str, location="json")
    parser.add_argument("contract_company", type=str, location="json")
    parser.add_argument("company_id", type=int, location="json")
    parser.add_argument('project_item_ids', type=int, action='append')
    parser.add_argument("warehouse_code", type=str, location="json")
    args = parser.parse_args()


    from backend_model.table_production import WorkOrder, WorkOrderItem1
    from backend_lib.lib_common import LibCommon

    project_items = db.session.query(
        ProjectItem
    ).filter(
        ProjectItem.id.in_(args['project_item_ids'])
    ).all()
    data = {'fk_company_id': args['company_id']}
    LibCommon.get_item_number(data, 'number', WorkOrder, WorkOrder.number, '/produce/work-order')
    new_order = WorkOrder()
    new_order.number = data['number']
    new_order.target_date =  datetime.now()
    # new_order.client_company = args['contract_company']
    new_order.department = args['department']
    new_order.manager = args['manager']
    new_order.fk_company_id = args['company_id']
    db.session.add(new_order)
    db.session.commit()

    for project_item in project_items:
        new_order_item = WorkOrderItem1()
        new_order_item.item_code = project_item.item_code
        new_order_item.required_quantity = project_item.quantity
        new_order_item.bom_yn = project_item.item.bom_yn
        if project_item.shipment_order is not None:
            new_order_item.order_number = project_item.shipment_order.order_number
        # new_order_item.request_delivery_date =
        new_order_item.warehouse_code = args['warehouse_code']
        new_order_item.unproduced_quantity = project_item.quantity
        new_order_item.client_company = args['contract_company']
        new_order_item.fk_work_order_id = new_order.id
        new_order_item.fk_project_item_id = project_item.id
        new_order_item.fk_project_management_id = args['project_management_id']
        project_item.fk_work_order_id = new_order.id
        db.session.add(new_order_item)
    db.session.commit()
    return make_response('success', 200)


@app.route("/api/mes/v1/project/registration-to-shipment-order", methods=["POST"])
def registration_to_shipment_order():
    parser = reqparse.RequestParser()
    parser.add_argument("project_management_id", type=int, location="json")
    parser.add_argument("department", type=str, location="json")
    parser.add_argument("manager", type=str, location="json")
    parser.add_argument("order_company", type=str, location="json")
    parser.add_argument("company_id", type=int, location="json")
    parser.add_argument('project_item_ids', type=int, action='append')
    parser.add_argument("warehouse_code", type=str, location="json")
    args = parser.parse_args()

    from backend_model.table_base import BaseCode
    from backend_lib.lib_common import LibCommon
    from backend_model.table_shipment import ShipmentOrder, ShipmentOrderItem
    add_count = 0
    vat_code = db.session.query(BaseCode).filter(BaseCode.code_name == '부가세구분').first()
    vat_type = None
    if vat_code and len(vat_code.items) > 0:
        vat_type = vat_code.items[0].code_name
    order_code = db.session.query(BaseCode).filter(BaseCode.code_name == '수주구분').first()
    order_type = None
    if order_code and len(order_code.items) > 0:
        order_type = order_code.items[0].code_name
    payment_code = db.session.query(BaseCode).filter(BaseCode.code_name == '결재조건').first()
    payment_type = None
    if payment_code and len(payment_code.items) > 0:
        payment_type = payment_code.items[0].code_name
        
    project_items = db.session.query(
        ProjectItem
    ).filter(
        ProjectItem.id.in_(args['project_item_ids'])
    ).all()
    data = {'fk_company_id': args['company_id']}
    LibCommon.get_item_number(data, 'order_number', ShipmentOrder, ShipmentOrder.order_number, '/shipment/order')
    new_order = ShipmentOrder()
    new_order.order_number = data['order_number']
    new_order.order_date =  datetime.now()
    new_order.client_company = args['order_company']
    new_order.order_department = args['department']
    new_order.order_manager = args['manager']
    new_order.order_type = order_type
    new_order.vat_type = vat_type
    new_order.payment_terms = payment_type
    new_order.fk_project_management_id = args['project_management_id']
    # new_order.supply_price = 
    # new_order.vat =
    # new_order.total_price =
    new_order.fk_company_id = args['company_id']
    db.session.add(new_order)
    db.session.commit()
    add_count += 1
    for project_item in project_items:
        new_order_item = ShipmentOrderItem()
        new_order_item.item_code = project_item.item_code
        new_order_item.order_quantity = project_item.quantity
        # new_order_item.assign_quantity = 
        # new_order_item.unit_price =
        # new_order_item.supply_price = 
        new_order_item.warehouse_code = args['warehouse_code']
        new_order_item.not_shipped = project_item.quantity
        new_order_item.fk_order_id = new_order.id
        new_order_item.fk_project_item_id = project_item.id
        new_order_item.fk_project_management_id = args['project_management_id']
        project_item.not_ordered = 0
        project_item.fk_shipment_order_id = new_order.id
        db.session.add(new_order_item)
        add_count += 1
    db.session.commit()
    if add_count <= 0:
        from backend_lib.lib_exception import MesException
        MesException.raise_not_add_item()
    return make_response('success', 200)

    # parser = reqparse.RequestParser()

@app.route("/api/mes/v1/project/excution-item-to-order", methods=["POST"])
def excution_item_to_order():
    parser = reqparse.RequestParser()
    parser.add_argument("id", type=int, location="json")
    parser.add_argument("department", type=str, location="json")
    parser.add_argument("manager", type=str, location="json")
    parser.add_argument("company_id", type=int, location="json")
    parser.add_argument('excution_plan_ids', type=int, action='append')
    args = parser.parse_args()

    from backend_model.table_base import BaseCode
    vat_code = db.session.query(BaseCode).filter(BaseCode.code_name == '부가세구분').first()
    vat_type = None
    if vat_code and len(vat_code.items) > 0:
        vat_type = vat_code.items[0].code_name
    order_code = db.session.query(BaseCode).filter(BaseCode.code_name == '발주구분').first()
    order_type = None
    if order_code and len(order_code.items) > 0:
        order_type = order_code.items[0].code_name
    payment_code = db.session.query(BaseCode).filter(BaseCode.code_name == '결재조건').first()
    payment_type = None
    if payment_code and len(payment_code.items) > 0:
        payment_type = payment_code.items[0].code_name

    from backend_lib.lib_util import LibUtil
    from backend_lib.lib_common import LibCommon
    from backend_model.table_purchase import PurchaseOrderItem, PurchaseOrder
    add_count = 0
    valid_items = {}

    plan_items = db.session.query(
        ProjectExcutionPlanItem
    ).filter(
        ProjectExcutionPlanItem.id.in_(args['excution_plan_ids'])
    ).all()

    for plan_item in plan_items:
        if plan_item.not_excution_plan_quantity <= 0:
            continue
        if not plan_item.main_supplier:
            continue
        fk_project_management_id = None
        project_manager = None
        if plan_item.project_excution_plan is not None:
                fk_project_management_id = plan_item.project_excution_plan.fk_project_management_id
                project_manager = plan_item.project_excution_plan.project_management.project_manager
        if plan_item.order_date_modify:
            order_type = '납기변경발주'
        if plan_item.main_supplier in valid_items:
            new_order = valid_items[plan_item.main_supplier]
            price = LibUtil.calculate_price(vat_type, plan_item.not_excution_plan_quantity, plan_item.purchase_unit_price)
            new_order_item = PurchaseOrderItem()
            new_order_item.item_code = plan_item.item_code
            new_order_item.order_quantity = plan_item.not_excution_plan_quantity
            new_order_item.expect_unit_price = plan_item.unit_price #예정단가 할당
            new_order_item.unit_price = plan_item.purchase_unit_price #발주 단가 할당
            new_order_item.supply_price = plan_item.purchase_supply_price #발주 공급가 할당
            new_order_item.request_delivery_date = plan_item.delivery_date
            new_order_item.not_shipped = plan_item.not_excution_plan_quantity
            new_order_item.fk_project_management_id = fk_project_management_id
            new_order_item.warehouse_code = plan_item.warehouse_code
            new_order_item.closing_yn = 0
            new_order_item.fk_excution_plan_item_id = plan_item.id
            new_order_item.fk_purchase_order_id = new_order.id
            if plan_item.item.import_check:
                new_order_item.check_yn = 0
            else:
                new_order_item.check_yn = 1
            db.session.add(new_order_item)
            
            new_order.supply_price += price['supply_price']
            new_order.vat += price['vat']
            new_order.total_price += price['total_price']
            db.session.commit()
        else:
            data = {
                'fk_company_id': args['company_id']
            }
            LibCommon.get_item_number(data, 'order_number', PurchaseOrder, PurchaseOrder.order_number, '/purchase/order-plan')
            manager_email = None
            if project_manager:
                from backend_model.table_base import BaseEmployee
                employee = db.session.query(BaseEmployee).filter(BaseEmployee.emp_name == project_manager).first()
                if employee:
                    manager_email = employee.emp_email

            price = LibUtil.calculate_price(vat_type, plan_item.not_excution_plan_quantity, plan_item.purchase_unit_price)
            new_order = PurchaseOrder()
            new_order.order_number = data['order_number']
            new_order.fk_project_management_id = fk_project_management_id
            new_order.order_date = datetime.now()
            new_order.order_department = args['department']
            new_order.order_manager = args['manager']
            new_order.order_manager_email = manager_email
            new_order.client_manager = '관리자'
            new_order.client_manager_email = 'souljdh0736@gmail.com'
            new_order.client_company = plan_item.main_supplier
            new_order.fk_company_id = args['company_id']
            new_order.vat_type = vat_type
            new_order.order_type = order_type
            new_order.payment_terms = payment_type
            new_order.supply_price = price['supply_price']
            new_order.vat = price['vat']
            new_order.total_price = price['total_price']
            db.session.add(new_order)
            db.session.commit()

            new_order_item = PurchaseOrderItem()
            new_order_item.item_code = plan_item.item_code
            new_order_item.order_quantity = plan_item.not_excution_plan_quantity
            new_order_item.expect_unit_price = plan_item.unit_price #예정단가 할당
            new_order_item.unit_price = plan_item.purchase_unit_price #발주 단가 할당
            new_order_item.supply_price = plan_item.purchase_supply_price #발주 공급가 할당
            new_order_item.request_delivery_date = plan_item.delivery_date
            new_order_item.not_shipped = plan_item.not_excution_plan_quantity
            new_order_item.fk_project_management_id = fk_project_management_id
            new_order_item.warehouse_code = plan_item.warehouse_code
            new_order_item.closing_yn = 0
            new_order_item.fk_excution_plan_item_id = plan_item.id
            new_order_item.fk_purchase_order_id = new_order.id
            if plan_item.item.import_check:
                new_order_item.check_yn = 0
            else:
                new_order_item.check_yn = 1
            db.session.add(new_order_item)

            valid_items[plan_item.main_supplier] = new_order
        db.session.commit()
        LibProjectExcutionPlanItem.update_not_excution_plan_quantity_quantity(plan_item.id)
        add_count += 1
    if add_count <= 0:
        from backend_lib.lib_exception import MesException
        MesException.raise_not_add_item()
    return make_response('success', 200)

@app.route("/api/mes/v1/project/excution-subcontract-to-order", methods=["POST"])
def excution_subcontract_to_order():
    parser = reqparse.RequestParser()
    parser.add_argument("id", type=int, location="json")
    parser.add_argument("department", type=str, location="json")
    parser.add_argument("manager", type=str, location="json")
    parser.add_argument("company_id", type=int, location="json")
    parser.add_argument('excution_plan_ids', type=int, action='append')
    args = parser.parse_args()

    if args['manager']:
        from backend_model.table_base import BaseEmployee
        employee = db.session.query(BaseEmployee).filter(BaseEmployee.emp_name == args['manager']).first()
        if employee:
            manager_email = employee.emp_email

    from backend_model.table_base import BaseCode
    vat_code = db.session.query(BaseCode).filter(BaseCode.code_name == '부가세구분').first()
    vat_type = None
    if vat_code and len(vat_code.items) > 0:
        vat_type = vat_code.items[0].code_name
    order_code = db.session.query(BaseCode).filter(BaseCode.code_name == '발주구분').first()
    order_type = None
    if order_code and len(order_code.items) > 0:
        order_type = order_code.items[0].code_name
    payment_code = db.session.query(BaseCode).filter(BaseCode.code_name == '결재조건').first()
    payment_type = None
    if payment_code and len(payment_code.items) > 0:
        payment_type = payment_code.items[0].code_name

    from backend_lib.lib_util import LibUtil
    from backend_lib.lib_common import LibCommon
    from backend_model.table_purchase import PurchaseOrderItem, PurchaseOrder
    from backend_lib.lib_project import LibProjectExcutionPlanSubcontract
    add_count = 0
    valid_items = {}

    plan_subcontracts = db.session.query(
        ProjectExcutionPlanSubcontract
    ).filter(
        ProjectExcutionPlanSubcontract.id.in_(args['excution_plan_ids'])
    ).all()

    for plan_subcontract in plan_subcontracts:
        if plan_subcontract.closing_yn:
            continue
        if not plan_subcontract.subcontract_company:
            continue
        fk_project_management_id = None
        if plan_subcontract.project_excution_plan is not None:
                fk_project_management_id = plan_subcontract.project_excution_plan.fk_project_management_id
        if plan_subcontract.subcontract_company in valid_items:
            price = LibUtil.calculate_price(vat_type, 1, plan_subcontract.purchase_unit_price)
            new_order = valid_items[plan_subcontract.subcontract_company]
            new_order_item = PurchaseOrderItem()
            new_order_item.item_code = plan_subcontract.subcontract_name
            new_order_item.order_quantity = 1
            new_order_item.expect_unit_price = plan_subcontract.expect_amount #예정단가 할당
            new_order_item.unit_price = plan_subcontract.purchase_unit_price #발주 단가 할당
            new_order_item.supply_price = plan_subcontract.purchase_unit_price #발주 공급가 할당
            new_order_item.not_shipped = 1
            new_order_item.fk_project_management_id = fk_project_management_id
            # new_order_item.warehouse_code = plan_item.warehouse_code
            new_order_item.closing_yn = 0
            new_order_item.fk_excution_plan_subcontract_id = plan_subcontract.id
            new_order_item.fk_purchase_order_id = new_order.id
            new_order_item.check_yn = 1
            # if plan_item.item.import_check:
            #     new_order_item.check_yn = 0
            # else:
            #     new_order_item.check_yn = 1
            db.session.add(new_order_item)
            
            new_order.supply_price += price['supply_price']
            new_order.vat += price['vat']
            new_order.total_price += price['total_price']
            db.session.commit()
        else:
            data = {
                'fk_company_id': args['company_id']
            }
            LibCommon.get_item_number(data, 'order_number', PurchaseOrder, PurchaseOrder.order_number, '/purchase/order-plan')

            price = LibUtil.calculate_price(vat_type, 1, plan_subcontract.purchase_unit_price)
            new_order = PurchaseOrder()
            new_order.order_number = data['order_number']
            new_order.fk_project_management_id = fk_project_management_id
            new_order.order_date = datetime.now()
            new_order.order_department = args['department']
            new_order.order_manager = args['manager']
            new_order.order_manager_email = manager_email
            new_order.client_manager = '관리자'
            new_order.client_manager_email = 'souljdh0736@gmail.com'
            new_order.client_company = plan_subcontract.subcontract_company
            new_order.fk_company_id = args['company_id']
            new_order.vat_type = vat_type
            new_order.order_type = order_type
            new_order.payment_terms = payment_type
            new_order.supply_price = price['supply_price']
            new_order.vat = price['vat']
            new_order.total_price = price['total_price']
            db.session.add(new_order)
            db.session.commit()

            new_order_item = PurchaseOrderItem()
            new_order_item.item_code = plan_subcontract.subcontract_name
            new_order_item.order_quantity = 1
            new_order_item.expect_unit_price = plan_subcontract.expect_amount #예정단가 할당
            new_order_item.unit_price = plan_subcontract.purchase_unit_price #발주 단가 할당
            new_order_item.supply_price = plan_subcontract.purchase_unit_price #발주 공급가 할당
            new_order_item.not_shipped = 1
            new_order_item.fk_project_management_id = fk_project_management_id
            # new_order_item.warehouse_code = plan_item.warehouse_code
            new_order_item.closing_yn = 0
            new_order_item.fk_excution_plan_subcontract_id = plan_subcontract.id
            new_order_item.fk_purchase_order_id = new_order.id
            new_order_item.check_yn = 1
            # if plan_item.item.import_check:
            #     new_order_item.check_yn = 0
            # else:
            #     new_order_item.check_yn = 1
            db.session.add(new_order_item)

            valid_items[plan_subcontract.subcontract_company] = new_order
        db.session.commit()
        LibProjectExcutionPlanSubcontract.update_closing(plan_subcontract.id)
        add_count += 1
    if add_count <= 0:
        from backend_lib.lib_exception import MesException
        MesException.raise_not_add_item()
    return make_response('success', 200)


@app.route("/api/mes/v1/project/monitoring", methods=["POST"])
def project_monitoring():
    parser = reqparse.RequestParser()
    parser.add_argument("page", type=int, location="json")
    args = parser.parse_args()

    num_in_page = 6
    offset = (args['page'] - 1) * num_in_page
    projects = db.session.query(
        ProjectManagement
    ).filter(
        ProjectManagement.total_progress > 0
    ).filter(
        ProjectManagement.total_progress < 100
    ).order_by(
        ProjectManagement.contract_begin_date.asc()
    ).offset(
        offset
    ).limit(
        num_in_page
    ).all()
    data = []
    for project in projects:
        sum_of_quantity = db.session.query(
            db.func.sum(ProjectItem.quantity)
        ).filter(
            ProjectItem.fk_project_management_id == project.id
        ).first()[0]
        if sum_of_quantity is None:
            sum_of_quantity = 0

        contract_end_date = ""
        if project.contract_end_date is not None:
            contract_end_date = project.contract_end_date.strftime("%Y-%m-%d")

        data.append({
            'id': project.id,
            'project_number': project.project_number,
            'project_name': project.project_name,
            'total_progress': project.total_progress,
            'contract_end_date': contract_end_date,
            'sum_of_quantity': str(sum_of_quantity) + " EA"
        })

    total_count = db.session.query(
        ProjectManagement
    ).filter(
        ProjectManagement.total_progress > 0
    ).filter(
        ProjectManagement.total_progress < 100
    ).count()
    total_page = 1
    if total_count and total_count > 0:
        total_page = total_count / 6
        if total_count % 6 > 0:
            total_page += 1

    result = {
        'total': total_page,
        'data': data
    }
    return make_response(jsonify(result), 200)

@app.route("/api/mes/v1/project/business/excel/upload", methods=["POST"])
def project_business_upload():
    from backend.api_common import validate_token
    from backend import check_token, check_permission, get_user
    from backend_lib.lib_excel import import_excel, ImportExcelException
    from backend_model.table_project import ProjectBusiness
    from openpyxl import load_workbook, Workbook
    import tempfile
    
    token = request.headers.get('token')
    if token is None:
        return make_response('Not Authorized', 410)

    user = validate_token(token)
    exc = check_permission(user, method='EXCEL')
    if exc is not None:
        return make_response(exc.description, exc.code)

    f = request.files['file']
    if f is None:
        return make_response('require parameter is missing', 400)

    filepath, ext = f.filename.rsplit('.', 1)
    if ext not in ['xlsx']:
        return make_response('invalid file extension', 400)

    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False)

        wb = load_workbook(f)
        ws = wb.active
        
        new_wb = Workbook()
        new_ws = new_wb.active
        
        field_mapping = {
            '공고번호': 'business_number',
            '공고명': 'business_name',
            '발주처': 'client_company',
            '종목': 'classification',
            # '지역': 'location',
            '참가등록마감': 'registration_closing_date',
            '투찰마감': 'bid_closing_date',
            '기초금액': 'business_amount'
        }
        header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]

        column_indices = {}
        for field_name in field_mapping.keys():
            found = False
            for idx, cell_value in enumerate(header_row[1:], 1):  # No 컬럼 제외하고 시작
                if cell_value:
                    # 셀 값과 필드명 정리 (개행, 공백, 특수문자 제거)
                    clean_cell = cell_value.replace('\n', '').replace(' ', '').replace('(', '').replace(')', '')
                    clean_field = field_name.replace(' ', '')
                    
                    if clean_cell.startswith(clean_field):
                        column_indices[field_name] = idx
                        found = True
                        break
                        
            if not found:
                print(f"Warning: '{field_name}' 컬럼을 찾을 수 없습니다.")
     
        headers = list(field_mapping.values())
        new_ws.append(headers)
        
        # 5번째 행 이후부터 데이터 시작
        for idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), 6):
            try:
                new_row = []
   
                for field_name in field_mapping.keys():
                    col_idx = column_indices[field_name]
                    value = row[col_idx] if col_idx < len(row) else None
                    
                    if field_name in ['투찰마감', '참가등록마감'] and value:
                        try:
                            date_str = value.replace('(', '').replace(')', '').strip()
                            date_obj = datetime.strptime('20' + date_str, '%Y.%m.%d %H:%M')
                            value = date_obj
                        except Exception as e:
                            print(f"날짜 변환 중 오류 ({field_name}): {e}, 값: {value}")
                            value = None
                    
                    new_row.append(value)
                
                if any(new_row):  # 빈 행 제외
                    new_ws.append(new_row)
                    
            except Exception as row_error:
                print(f"행 처리 중 오류 발생 ({idx}번째 행):", str(row_error))
                raise

        new_wb.save(temp_file.name)
        common_values = dict(
            fk_company_id=user.fk_company_id, 
            modify_manager=user.user_id, 
            modify_date=datetime.now(),
            business_type='입찰',
            business_department='경영지원팀',
            business_manager='오선아',
            business_status='신규',
            business_progress='진행',
            business_important='일반'
        )

        with open(temp_file.name, 'rb') as converted_file:
            import_excel(
                table=ProjectBusiness, 
                common_values=common_values, 
                file=converted_file
            )
            
        return make_response('success', 200)
        
    except ImportExcelException as e:
        print(e)
        result = {
            'message': e.message if e.message else '데이터를 불러오는 중 에러가 발생하였습니다',
            'description': str(e)
        }
        return make_response(jsonify(result), 400)
        
    except Exception as e:
        print("상세 오류:", str(e))
        return make_response(f'엑셀 파일 처리 중 오류가 발생했습니다: {str(e)}', 400)
        
    finally:
        # 임시 파일 삭제
        if 'temp_file' in locals():
            temp_file.close()
            os.unlink(temp_file.name)


def post_project_management_preprocessor(data=None, **kw):
    menu = db.session.query(SetupMenu).filter(SetupMenu.path == '/project/registration').first()
    if menu is None:
        from backend_lib.lib_exception import MesException
        MesException.raise_not_exist_mapping_menu()

    start, end = menu.get_range()

    count = db.session.query(ProjectManagement)\
        .filter(ProjectManagement.created >= start)\
        .filter(ProjectManagement.created < end)\
        .count()

    data['project_number'] = menu.menu_number_format(count=count, default_initial='PROJT')


manager.create_api(ProjectManagement,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_management',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token, LibProjectManagement.post_preprocessor],
                       'PATCH_SINGLE': [check_token, LibProjectManagement.patch_single_preprocessor],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectManagement.delete_single_preprocessor]
                   },
                   postprocessors={
                       'POST': [check_token, LibProjectManagement.post_postprocessors],
                       'PATCH_SINGLE': [check_token]
                   })


manager.create_api(ProjectItem,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_item',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectSchedule,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_schedule',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token, LibProjectSchedule.update_project_progress],
                       'PATCH_SINGLE': [check_token, LibProjectSchedule.update_project_progress],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectSchedule.update_project_progress]
                   })

manager.create_api(ProjectNotice,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_notice',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token]
                   })

manager.create_api(ProjectBusiness,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_business',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token, LibProjectBusiness.post_preprocessor],
                       'PATCH_SINGLE': [check_token, LibProjectBusiness.patch_single_preprocessor],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectBusiness.delete_single_preprocessors]
                   },
                   postprocessors={
                       'POST': [check_token, LibProjectBusiness.post_postprocessors]
                   })

manager.create_api(ProjectBusinessNote,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_business_note',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectBusinessProgress,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_business_progress',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectParticipant,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_participant',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectCompany,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_company',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectDocument,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_document',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectReceipt,
                   url_prefix='/api/mes/v1/project',
                   collection_name='project_receipt',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })


manager.create_api(ProjectExcutionPlan,
                   url_prefix='/api/mes/v1/project',
                   collection_name='excution_plan',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token, LibProjectExcutionPlan.post_preprocessor],
                       'PATCH_SINGLE': [check_token, LibProjectExcutionPlan.patch_single_preprocessor],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectExcutionPlan.delete_single_preprocessor]
                   },
                   postprocessors={
                       'POST': [check_token, LibProjectExcutionPlan.post_postprocessors],
                       'PATCH_SINGLE': [check_token]
                   })


manager.create_api(ProjectExcutionPlanItem,
                   url_prefix='/api/mes/v1/project',
                   collection_name='excution_plan_item',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   },
                   postprocessors={
                       
                   })


manager.create_api(ProjectExcutionPlanSubcontract,
                   url_prefix='/api/mes/v1/project',
                   collection_name='excution_plan_subcontract',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })


manager.create_api(ProjectExcutionPlanExpense,
                   url_prefix='/api/mes/v1/project',
                   collection_name='excution_plan_expenxe',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectBusinessTripLog,
                   url_prefix='/api/mes/v1/project',
                   collection_name='business_trip_log',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectHappyCall,
                   url_prefix='/api/mes/v1/project',
                   collection_name='happy_call',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectDailyLog,
                   url_prefix='/api/mes/v1/project',
                   collection_name='daily_log',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectCostLog,
                   url_prefix='/api/mes/v1/project',
                   collection_name='cost_log',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectCompletion,
                   url_prefix='/api/mes/v1/project',
                   collection_name='completion',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,  
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectCustomerInformation,
                   url_prefix='/api/mes/v1/project',
                   collection_name='customer_information',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectCustomerHistory,
                   url_prefix='/api/mes/v1/project',
                   collection_name='customer_history',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })   

manager.create_api(ProjectBusinessQuote,
                   url_prefix='/api/mes/v1/project',
                   collection_name='business_quote',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectBusinessQuote.delete_single_preprocessor]
                   })

manager.create_api(ProjectBusinessCost,
                   url_prefix='/api/mes/v1/project',
                   collection_name='business_cost',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token]
                   })

manager.create_api(ProjectBusinessBasic,
                   url_prefix='/api/mes/v1/project',
                   collection_name='business_basic',
                   methods=['GET', 'DELETE', 'PATCH', 'POST'],
                   allow_patch_many=True,
                   results_per_page=0,
                   max_results_per_page=100000000,
                   preprocessors={
                       'POST': [check_token],
                       'PATCH_SINGLE': [check_token],
                       'GET_SINGLE': [check_token],
                       'GET_MANY': [check_token],
                       'DELETE_SINGLE': [check_token, LibProjectBusinessBasic.delete_single_preprocessor]
                   })
