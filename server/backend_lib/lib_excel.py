# -*- coding: utf-8 -*-
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from backend_model.database import DBManager
from sqlalchemy import exc
from datetime import datetime
import json


class ImportExcelException(Exception):
    def __init__(self, description, message=None):
        self.description = description
        self.message = message

    def __str__(self):
        return 'import excel exception: {}'.format(self.description)


def response_to_excel(raw):
    data = json.loads(raw)
    objects = data['objects']

    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet'

    for row_index, row in enumerate(objects):
        col_index = 1
        for col_name in row:
            ws.cell(column=col_index, row=row_index + 1, value=row[col_name])
            col_index += 1

    data = BytesIO()
    wb.save(data)
    return data


def sample_excel(table=None):
    """
    :param table: Model
    :return: BytesIO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet'
    header_style = PatternFill(start_color='00FFFF99', end_color='00FFFF99', fill_type='solid')

    columns = table.excel_import_columns(table)
    for idx, col in enumerate(columns['output']):
        column = col
        if type(column) is tuple:
            col, sec_model, input_key, link_key = col
            column = getattr(sec_model, input_key)
        coltype = str(column.expression.type)
        def_value = ''
        letter = get_column_letter(idx + 1)
        ws.column_dimensions[letter].width = 20
        if coltype.startswith('DATETIME'):
            def_value = datetime.now()
        elif coltype.startswith('INTEGER'):
            def_value = 0
        elif coltype.startswith('BOOLEAN'):
            def_value = 'Y/N'

        h_cell = ws.cell(column=idx + 1, row=1, value=column.expression.comment)
        h_cell.fill = header_style
        ws.cell(column=idx + 1, row=2, value=def_value)

    data = BytesIO()
    wb.save(data)
    return data


def import_excel(table=None, common_values=None, file=None):
    """
    :param table: 테이블 인스턴스 (ex. BaseItem)
    :param common_values: 모든 행에 공통으로 들어갈 공통 값 (ex. dict(fk_company_id=1))
    :param file: 업로드된 엑셀 파일
    :return: None
    """
    if table is None or file is None:
        raise ImportExcelException(description="Missing required parameter")

    db = DBManager.db

    temporary_options = dict()
    columns = table.excel_import_columns(table)
    for column in columns['output']:
        if type(column) is tuple:
            col, sec_model, input_key, link_key = column
            # common_values.fk_company_id
            temporary_options[col.key] = dict()
            for row in sec_model.query.all():
                temporary_options[col.key][getattr(row, input_key)] = getattr(row, link_key)

    wb = load_workbook(filename=file)
    ws = wb.active

    for idx, row in enumerate(ws):
        if idx == 0:
            continue

        record = dict()
        for i, column in enumerate(columns['output']):
            if type(column) is tuple:
                col, sec_model, input_key, link_key = column
                if row[i].value in temporary_options[col.key]:
                    ref = temporary_options[col.key][row[i].value]
                    record[col.key] = ref
            else:
                coltype = str(column.type)
                if coltype.startswith('BOOLEAN'):
                    if row[i].value == 'Y':
                        row[i].value = True
                    else:
                        row[i].value = False
                record[column.key] = row[i].value

        if common_values is not None:
            for key, value in common_values.items():
                record[key] = value

        q = db.session.query(table)
        for key in columns['key']:
            q = q.filter(key == record[key.key])
            
        row = q.first()

        if row is None:
            row = table()
            for key in record:
                setattr(row, key, record[key])
            db.session.add(row)
        else:
            q.update(record)

    try:
        db.session.commit()
        print('inserted {} rows'.format(idx + 1))
    except exc.IntegrityError as e:
        db.session.rollback()
        raise ImportExcelException(description=str(e), message='중복된 키를 가진 데이터가 있습니다')
    except exc.SQLAlchemyError as e:
        db.session.rollback()
        raise ImportExcelException(description=str(e))


def parse_excel(table=None, common_values=None, file=None):
    """
    :param table: 테이블 인스턴스 (ex. BaseItem)
    :param common_values: 모든 행에 공통으로 들어갈 공통 값 (ex. dict(fk_company_id=1))
    :param file: 업로드된 엑셀 파일
    :return: None
    """
    if table is None or file is None:
        raise ImportExcelException(description="Missing required parameter")

    temporary_options = dict()
    columns = table.excel_import_columns(table)
    for column in columns['output']:
        if type(column) is tuple:
            col, sec_model, input_key, link_key = column
            # common_values.fk_company_id
            temporary_options[col.key] = dict()
            for row in sec_model.query.all():
                temporary_options[col.key][getattr(row, input_key)] = getattr(row, link_key)

    wb = load_workbook(filename=file)
    ws = wb.active

    records = []
    for idx, row in enumerate(ws):
        if idx == 0:
            continue

        record = dict()
        for i, column in enumerate(columns['output']):
            if type(column) is tuple:
                col, sec_model, input_key, link_key = column
                if row[i].value in temporary_options[col.key]:
                    ref = temporary_options[col.key][row[i].value]
                    record[col.key] = ref
            else:
                coltype = str(column.type)
                if coltype.startswith('BOOLEAN'):
                    if row[i].value == 'Y':
                        row[i].value = True
                    else:
                        row[i].value = False
                record[column.key] = row[i].value

        if common_values is not None:
            for key, value in common_values.items():
                record[key] = value
        records.append(record)
    return records


base_bom_excel_headers = [
    dict(key='item_code', name='품목코드', type='STRING'),
    dict(key='lrate', name='L/Rate', type='INTEGER'),
    dict(key='requirement', name='소요량', type='INTEGER'),
    dict(key='consume_yn', name='자재소모 여부', type='BOOLEAN')
]


def sample_excel_bom():
    """
    :param table: Model
    :return: BytesIO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet'
    header_style = PatternFill(start_color='00FFFF99', end_color='00FFFF99', fill_type='solid')

    for idx, col in enumerate(base_bom_excel_headers):
        def_value = ''
        letter = get_column_letter(idx + 1)
        ws.column_dimensions[letter].width = 20
        if col['type'].startswith('DATETIME'):
            def_value = datetime.now()
        elif col['type'].startswith('INTEGER'):
            def_value = 0
        elif col['type'].startswith('BOOLEAN'):
            def_value = 'Y/N'

        h_cell = ws.cell(column=idx + 1, row=1, value=col['name'])
        h_cell.fill = header_style
        ws.cell(column=idx + 1, row=2, value=def_value)

    data = BytesIO()
    wb.save(data)
    return data


def import_excel_bom(common_values=None, file=None):
    from backend_model.table_base import BaseBOM, BaseBOMLink, BaseItem

    db = DBManager.db

    wb = load_workbook(filename=file)
    ws = wb.active

    for row_index, row in enumerate(ws):
        if row_index == 0:
            continue

        record = dict()
        for col_index, col in enumerate(base_bom_excel_headers):
            record[base_bom_excel_headers[col_index]['key']] = row[col_index].value
        item = BaseItem.query.filter_by(item_code=record['item_code']).first()
        if not item:
            print(f'import bom: {record["item_code"]} -> not found')
            raise ImportExcelException(description="Missing required parameter")

        bom = BaseBOM.query.filter_by(item_id=item.id).filter_by(fk_company_id=common_values['fk_company_id']).first()
        if not bom:
            bom = BaseBOM()
            bom.item_id = item.id
            bom.bom_depth = 1
            bom.fk_company_id = common_values['fk_company_id']
            db.session.add(bom)
            db.session.flush()

        link = BaseBOMLink.query\
            .filter_by(root_id=common_values['root_id'])\
            .filter_by(child_id=bom.id)\
            .filter_by(parent_id=common_values['root_id'])\
            .first()

        if not link:
            link = BaseBOMLink()
            link.root_id = common_values['root_id']
            link.child_id = bom.id
            link.parent_id = common_values['root_id']
            link.consume_yn = True if record['consume_yn'] == 'Y' else False
            link.requirement = record['requirement']
            link.lrate = record['lrate']
            db.session.add(link)
        else:
            link.consume_yn = True if record['consume_yn'] == 'Y' else False
            link.requirement = record['requirement']
            link.lrate = record['lrate']
        db.session.commit()
